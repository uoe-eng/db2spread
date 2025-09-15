import sqlalchemy as sa
import openpyxl as px

from argparse import ArgumentParser
from configparser import ConfigParser
from dataclasses import dataclass
from sqlalchemy import engine_from_config
from typing import Iterable
@dataclass
class DB2Spread:
    engine: sa.Engine.engine
    models: Iterable = None
    workbook: px.Workbook = None
    column_handlers: dict = None
    limit: int = None

    def __post_init__(self):
        if self.models is None:
            self.models = []
        if self.workbook is None:
            self.workbook = px.Workbook()
        if self.column_handlers is None:
            self.column_handlers = {}

    def add_column_handler(self, column_name, handler):
        self.column_handlers[column_name] = handler

    def export(self, filename: str):
        for model in self.models:
            sheet = self.workbook.create_sheet(title=model.__tablename__)
            columns = list(model.__mapper__.attrs.keys())
            columns.extend(model.__mapper__.relationships.keys())
            sheet.append(columns)

            with self.engine.connect() as connection:
                result = connection.execute(sa.select(model).limit(self.limit) if self.limit else sa.select(model))
                for row in result:
                    processed_row = []
                    for col in columns:
                        value = getattr(row, col)
                        if col in self.column_handlers:
                            value = self.column_handlers[col](value)
                        processed_row.append(value)
                    sheet.append(processed_row)

        if 'Sheet' in self.workbook.sheetnames:
            std = self.workbook['Sheet']
            self.workbook.remove(std)
        self.workbook.save(filename)


if __name__ == "__main__":
    parser = ArgumentParser(description="Export database tables to Excel spreadsheet.")
    parser.add_argument("-l", "--limit", type=int, help="Limit number of rows per table.", default=None)
    parser.add_argument("-m", "--module", help="Module containing SQLAlchemy models.", required=True)
    parser.add_argument("-o", "--output", help="Output Excel file name.", default="output.xlsx")
    parser.add_argument("config", help="Path to the database configuration file.")
    args = parser.parse_args()

    config = ConfigParser()
    config.read(args.config)
    module = __import__(args.module, fromlist=['*'])
    models = [getattr(module, attr) for attr in dir(module) if hasattr(getattr(module, attr), '__tablename__')]
    engine = engine_from_config(config, prefix='sqlalchemy.')

    db2spread = DB2Spread(engine=engine, models=models, limit=args.limit)

    db2spread.export(args.output)
